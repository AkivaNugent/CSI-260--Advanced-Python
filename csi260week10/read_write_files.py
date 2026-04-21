import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

with open("data.csv", "r") as f:
    print(f.read())


df = pd.read_csv('data_no_dates.csv')

print(df)
print(df.corr()) #Outputs the correlation coefficients for the variables

#df.plot()

#plt.show()

df = pd.read_csv('data.csv')

#df.plot(kind = 'scatter', x = 'Duration', y = 'Calories')

#plt.show()

# Give the location of the file
path = "sales_report.xlsx"

'''
 To open the workbook
 workbook object is created
'''
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

cell_obj = sheet_obj['A1': 'B6']

for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)