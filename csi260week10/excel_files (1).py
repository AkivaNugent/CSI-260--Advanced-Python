import pandas as pd
import openpyxl

wb = openpyxl.load_workbook("sample.xlsx")

sheet = wb.active
'''
c = sheet['A3']
c.value = "New Data"

wb.save("sample.xlsx")
sheet = wb.active

data = (
    (1, 2, 3),
    (4, 5, 6)
)

for row in data:
    sheet.append(row)

wb.save('sample.xlsx')


# writing to the specified cell
sheet.cell(row=1, column=1).value = ' hello '

sheet.cell(row=2, column=2).value = ' everyone '

# set the height of the row
sheet.row_dimensions[1].height = 70

# set the width of the column
sheet.column_dimensions['B'].width = 20

# save the file
wb.save('sample.xlsx')


# writing to the cell of an Excel sheet
sheet['A10'] = 200
sheet['A11'] = 300
sheet['A12'] = 400
sheet['A13'] = 500
sheet['A14'] = 600

sheet['A15'] = '= SUM(A10:A14)'

# save the file
wb.save("sample.xlsx")


sheet.merge_cells('E2:H4')

sheet.cell(row=2, column=5).value = 'Twelve cells join together.'

# merge cell C6 and D6
sheet.merge_cells('E6:F6')

sheet.cell(row=6, column=5).value = 'Two merged cells.'

wb.save('sample.xlsx')
'''
#Charting
from openpyxl.chart import BarChart, Reference
df = pd.read_excel('sample.xlsx', sheet_name='chart')
# write o to 9 in 1st column of the active sheet
for i in range(10):
    sheet.append([i])

# create data for plotting
values = Reference(sheet, min_col=1, min_row=1,
                   max_col=1, max_row=10)

# Create object of BarChart class
chart = BarChart()

# adding data to the Bar chart object
chart.add_data(values)

# set the title of the chart
chart.title = " BAR-CHART "

# set the title of the x-axis
chart.x_axis.title = " X_AXIS "

# set the title of the y-axis
chart.y_axis.title = " Y_AXIS "

sheet.add_chart(chart, "E2")

# save the file
wb.save("sample.xlsx")