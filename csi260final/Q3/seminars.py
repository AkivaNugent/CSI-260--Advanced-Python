import openpyxl

wb = openpyxl.load_workbook("Seminars.xlsx")

def mgr_data():
    wb = openpyxl.load_workbook("Seminars.xlsx")

    calc_sheet = wb.create_sheet("calculations")

    calc_sheet['A1'] = 'Average'
    calc_sheet['B1'] = '=AVERAGEIF(Seminars!B2:B67, "Accounting", Seminars!J2:J67)'
    calc_sheet['B1'].number_format = '"$"#,##'

    calc_sheet['A2'] = 'Total Seminar Price per Person'
    calc_sheet['B2'] = '=SUM(Seminars!J2:J67)'
    calc_sheet['B2'].number_format = '"$"#,##'

    calc_sheet.column_dimensions['A'].width = 30
    wb.save("Seminars.xlsx")

mgr_data()


